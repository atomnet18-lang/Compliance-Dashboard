#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
convert_sharepoint_to_json.py  (v3 — รายละเอียดรายเดือนแยกคอลัมน์)
แปลงข้อมูลจาก SharePoint List (export .xlsx/.csv) -> data.json

คอลัมน์ที่ใช้ (1 แถว = 1 กิจกรรม):
  AID         : รหัสกิจกรรม เช่น 1-01   (KEY — ห้ามแก้)
  StartMonth  : เดือนเริ่ม   เช่น ม.ค.69
  EndMonth    : เดือนสิ้นสุด เช่น เม.ย.69
  Status      : ยังไม่เริ่ม | กำลังดำเนินการ | เสร็จสิ้น
  D_<เดือน>   : รายละเอียดงานของเดือนนั้น 15 คอลัมน์ เช่น D_ม.ค.69, D_ก.พ.69, ... D_มี.ค.70
                กรอกเฉพาะเดือนที่กิจกรรมดำเนินอยู่ (ในช่วง Start–End) ที่เหลือเว้นว่าง
  Note        : หมายเหตุภายใน (ไม่เผยแพร่)

วิธีใช้:
  python3 convert_sharepoint_to_json.py <list_export.xlsx|csv> [data.json]
"""
import sys, json, os, datetime

STATUS_MAP = {
    "ยังไม่เริ่ม": "none", "none": "none", "": "none",
    "กำลังดำเนินการ": "prog", "กำลังทำ": "prog", "prog": "prog",
    "เสร็จสิ้น": "done", "เสร็จ": "done", "done": "done",
}

def load_rows(path):
    ext = os.path.splitext(path)[1].lower()
    rows = []
    if ext in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        idx = {h: i for i, h in enumerate(header)}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or all(v is None for v in r):
                continue
            rows.append({h: (r[idx[h]] if idx[h] < len(r) else None) for h in header})
    else:
        import csv
        with open(path, encoding="utf-8-sig") as f:
            for d in csv.DictReader(f):
                rows.append(d)
    return rows

def main():
    if len(sys.argv) < 2:
        print("usage: python3 convert_sharepoint_to_json.py <list_export.xlsx|csv> [data.json]")
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "data.json"

    here = os.path.dirname(os.path.abspath(__file__))
    template = json.load(open(os.path.join(here, "data.json"), encoding="utf-8"))
    months = template["months"]
    midx = {m: i for i, m in enumerate(months)}
    # ชื่อคอลัมน์รายเดือน: D_<เดือน>
    month_cols = {m: f"D_{m}" for m in months}

    rows = load_rows(src)
    by_aid = {}
    for r in rows:
        aid = str(r.get("AID", "")).strip()
        if aid:
            by_aid[aid] = r

    PUBLIC = ("aid", "name", "resp", "indent")
    warnings = []
    for s in template["steps"]:
        for a in s["acts"]:
            for k in list(a.keys()):
                if k not in PUBLIC:
                    a.pop(k, None)
            a.setdefault("start", None); a.setdefault("end", None)
            a.setdefault("status", "none"); a.setdefault("details", {})
            row = by_aid.get(a["aid"])
            if not row:
                continue
            sm = str(row.get("StartMonth", "") or "").strip()
            em = str(row.get("EndMonth", "") or "").strip()
            st = str(row.get("Status", "") or "").strip()
            a["status"] = STATUS_MAP.get(st, "none")
            if sm and sm not in midx:
                warnings.append(f"{a['aid']}: StartMonth '{sm}' ไม่ตรงกับ months")
            if em and em not in midx:
                warnings.append(f"{a['aid']}: EndMonth '{em}' ไม่ตรงกับ months")
            s_i = midx.get(sm); e_i = midx.get(em)
            if s_i is not None and e_i is not None:
                if e_i < s_i:
                    s_i, e_i = e_i, s_i
                a["start"], a["end"] = s_i, e_i
            elif s_i is not None:
                a["start"] = a["end"] = s_i
            # รายละเอียดรายเดือน จาก 15 คอลัมน์
            details = {}
            for m, col in month_cols.items():
                val = row.get(col)
                if val is not None and str(val).strip():
                    details[str(midx[m])] = str(val).strip()
            a["details"] = details
            # เตือนถ้ามีรายละเอียดของเดือนที่อยู่นอกช่วง start-end
            if a["start"] is not None:
                for k in details:
                    ki = int(k)
                    if ki < a["start"] or ki > a["end"]:
                        warnings.append(f"{a['aid']}: มีรายละเอียดเดือน {months[ki]} ซึ่งอยู่นอกช่วงเวลา {months[a['start']]}–{months[a['end']]}")

    template["updated"] = datetime.datetime.now().strftime("%d/%m/") + str(datetime.datetime.now().year + 543)
    json.dump(template, open(out, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    done = sum(1 for s in template["steps"] for a in s["acts"] if a["status"] == "done")
    prog = sum(1 for s in template["steps"] for a in s["acts"] if a["status"] == "prog")
    ndet = sum(len(a["details"]) for s in template["steps"] for a in s["acts"])
    print(f"เขียน {out} แล้ว | เสร็จ {done} · กำลังทำ {prog} · รายละเอียดรายเดือน {ndet} รายการ")
    for w in warnings:
        print("  ⚠", w)

if __name__ == "__main__":
    main()
